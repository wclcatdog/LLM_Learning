from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side

from openpyxl import load_workbook
from os import listdir, path
from openpyxl.styles import PatternFill, Border, Side


def add_columns_and_style(file_path):
    # 加载Excel文件
    wb = load_workbook(file_path)

    # 遍历所有工作表
    for sheet in wb.worksheets:
        # 获取当前工作表的最后一列索引
        last_col_idx = sheet.max_column

        # 在最后一列的右侧插入两列
        sheet.insert_cols(last_col_idx + 1, 2)

        # 设置新插入的列的标题
        sheet.cell(row=1, column=last_col_idx + 1, value="复核结果")
        sheet.cell(row=1, column=last_col_idx + 2, value="复核人")

        # 设置标题的背景色
        sheet.cell(row=1, column=last_col_idx + 1).fill = PatternFill(start_color="FFFF00", end_color="FFFF00",
                                                                      fill_type="solid")
        sheet.cell(row=1, column=last_col_idx + 2).fill = PatternFill(start_color="FFFF00", end_color="FFFF00",
                                                                      fill_type="solid")

        # 设置整个工作表的边框




        for row in sheet.iter_rows():
            for cell in row:
                cell.border = Border(left=Side(border_style="thin", color="000000"),
                                     top=Side(border_style="thin", color="000000"),
                                     right=Side(border_style="thin", color="000000"),
                                     bottom=Side(border_style="thin", color="000000"))







                # 保存修改后的Excel文件
    wb.save(file_path)


def process_excel_files_in_folder(folder_path):
    # 遍历文件夹
    for filename in listdir(folder_path):
        # 检查文件扩展名是否为.xlsx
        if filename.endswith('.xlsx'):
            # 构建完整的文件路径
            file_path = path.join(folder_path, filename)
            print(file_path)
            # 对文件进行修改
            add_columns_and_style(file_path)
            print(f"Processed: {file_path}")

        # 使用示例


folder_path = '待复核文件夹'  # 替换为你的文件夹路径
process_excel_files_in_folder(folder_path)

