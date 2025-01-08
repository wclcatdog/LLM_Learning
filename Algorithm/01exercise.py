import os
import shutil

# 定义文件夹路径
folder1_path = 'C://Users//孙先生//Desktop//串码导入'
folder2_path = 'C://Users//孙先生//Desktop//未成功导入'

# 检查文件夹1是否为空
if len(os.listdir(folder1_path)) == 0:
    print("文件夹1为空，不执行操作。")
else:
    # 获取文件夹1中的所有文件
    files = os.listdir(folder1_path)

    # 遍历文件列表并移动文件到文件夹2
    for file_name in files:
        file_path = os.path.join(folder1_path, file_name)
        destination_path = os.path.join(folder2_path, file_name)
        shutil.move(file_path, destination_path)
        print(f"移动文件 {file_name} 到文件夹2成功。")


'-----------------------------------------------------------------------------'

import os
from openpyxl import Workbook

# 定义文件夹路径
folder1_path = 'E://工作//RPA//每日串码导入_已成功导入'
folder2_path = 'E://工作//RPA//每日串码导入_未成功导入'



# 获取文件夹1中的文件名
folder1_files = os.listdir(folder1_path)

# 获取文件夹2中的文件名
folder2_files = os.listdir(folder2_path)

# 创建一个新的 Excel 工作簿
workbook = Workbook()

# 获取默认的活动工作表
sheet = workbook.active

# 写入文件夹名称到第一行
sheet.cell(row=1, column=1).value = "已成功导入"
sheet.cell(row=1, column=2).value = "未成功导入"

# 写入文件名到第二行及以后的行
for index, file_name in enumerate(folder1_files, start=2):
    sheet.cell(row=index, column=1).value = file_name

for index, file_name in enumerate(folder2_files, start=2):
    sheet.cell(row=index, column=2).value = file_name

# 保存 Excel 文件
workbook.save('E://工作//RPA//每日串码导入_导入结果统计表.xlsx')